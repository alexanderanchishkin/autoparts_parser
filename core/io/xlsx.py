import os.path
from config import settings
import time

from openpyxl import load_workbook, Workbook
from core.models.part import Part
from core.utilities.value import has_cyrillic


def read_parts_from_xlsx(xlsx_file, max_amount=int(1e+6)):
    wb = load_workbook(xlsx_file, read_only=True)
    ws = wb.worksheets[0]
    start_row = 2 if not ws[1][0].value or has_cyrillic(ws[1][0].value) else 1
    parts = \
        [Part(row[0].value.split('#')[0], row[1].value)
         for row in ws.iter_rows(start_row, min(ws.max_row, max_amount), 1, 3) if row[0].value and row[1].value]
    return parts


def write_column_to_xlsx(xlsx_file, xlsx_table, time_moment):
    xlsx_file_path = settings.TEMP_XLSX_DIRECTORY + xlsx_file

    if not os.path.exists(xlsx_file_path):
        create_xlsx(xlsx_file_path, xlsx_table)

    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    add_new_column(ws, time_moment)
    while True:
        try:
            wb.save(xlsx_file_path)
            break
        except PermissionError:
            print('Закройте файл ', xlsx_file_path)
            time.sleep(3)


def write_parts_to_xlsx(xlsx_file, xlsx_table, parts, time_moment):
    return
    xlsx_file_path = settings.TEMP_XLSX_DIRECTORY + xlsx_file

    if not os.path.exists(xlsx_file_path):
        create_xlsx(xlsx_file_path, xlsx_table)
    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    last_column = get_column_by_index(ws.max_column)

    for part in parts.values():
        write_new_part(ws, part, last_column)

    while True:
        try:
            wb.save(xlsx_file_path)
            break
        except PermissionError:
            print('Закройте файл ', xlsx_file_path)
            time.sleep(3)


def write_new_part(table, part, last_column):
    row = table.max_row + 1
    table['A' + str(row)] = str(part.number)
    table['B' + str(row)] = part.model
    table['C' + str(row)] = part.title
    table[last_column + str(row)] = part.price


def add_new_column(table, time_moment):
    new_column_name = get_column_by_index(table.max_column + 1) + '1'
    table[new_column_name] = time_moment


def create_xlsx(xlsx_file_path, xlsx_table):
    wb = Workbook()
    ws = wb.active
    ws.title = xlsx_table

    column_names = ['Артикул', 'Бренд', 'Наименование']
    add_header_row(ws, column_names)

    wb.save(xlsx_file_path)


def add_header_row(table, column_names):
    for index in range(len(column_names)):
        cell_name = get_column_by_index(index + 1) + '1'
        table[cell_name] = column_names[index]


def get_column_by_index(index):  # col is 1 based
    excel_column = str()
    div = index
    while div:
        (div, mod) = divmod(div-1, 26)  # will return (x, 0 .. 25)
        excel_column = chr(mod + 65) + excel_column

    return excel_column


def merge_files(files, out_name):
    wb = Workbook()
    folder = 'files/results'
    temp_folder = os.path.join(folder, 'tmp')
    for file in files:
        filename = os.path.join(temp_folder, file)
        file_wb = load_workbook(filename, read_only=True)
        file_ws = file_wb.active
        wb.create_sheet(file_ws.title)
        for row in file_ws.iter_rows(1, file_ws.max_row, 1, 4):
            wb[file_ws.title].append((ceil.value for ceil in row))
    output_filename = f"{settings.time_moment_name}_{out_name}.xlsx"
    output_path = os.path.join(folder, output_filename)
    print(output_path)
    print(output_filename)
    wb.remove(wb.active)
    wb.save(output_path)
    return output_filename


def write_report(parts, out_name):
    wb = Workbook()
    folder = 'files/results'
    ws = wb.active

    ws.cell(1, 1).value = 'Артикул'
    ws.cell(1, 2).value = 'Бренд'
    ws.cell(1, 3).value = 'Минимум'
    ws.cell(1, 5).value = 'Средняя'
    ws.cell(1, 6).value = 'Максимум'
    ws.merge_cells('C1:D1')
    ws.merge_cells('F1:G1')
    ws.cell(2, 3).value = 'Цена'
    ws.cell(2, 4).value = 'Магазин'
    ws.cell(2, 6).value = 'Цена'
    ws.cell(2, 7).value = 'Магазин'

    for row, part in enumerate(parts.values()):
        ws.cell(row + 3, 1).value = part['article']
        ws.cell(row + 3, 2).value = part['brand']
        ws.cell(row + 3, 3).value = part['min_price']
        ws.cell(row + 3, 4).value = part['min_shop']
        ws.cell(row + 3, 5).value = part['avg_price']
        ws.cell(row + 3, 6).value = part['max_price']
        ws.cell(row + 3, 7).value = part['max_shop']

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

    for col, value in dims.items():
        ws.column_dimensions[chr(ord('A') + col - 1)].width = value * 1.5

    output_filename = f"{settings.time_moment_name}_{out_name}.xlsx"
    output_path = os.path.join(folder, output_filename)
    wb.save(output_path)
    return output_filename

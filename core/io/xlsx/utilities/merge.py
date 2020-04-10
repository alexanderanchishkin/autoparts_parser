import os.path

from openpyxl import Workbook, load_workbook

from config import settings


def merge_files(files, out_name):
    wb = Workbook()

    temp_folder = settings.TEMP_XLSX_DIRECTORY
    for file in files:
        filename = os.path.join(temp_folder, file)
        file_wb = load_workbook(filename, read_only=True)
        file_ws = file_wb.active
        wb.create_sheet(file_ws.title)

        for row in file_ws.iter_rows(1, file_ws.max_row, 1, 4):
            wb[file_ws.title].append((ceil.value for ceil in row))

    output_filename = f"{settings.time_moment_name}_{out_name}.xlsx"
    output_path = os.path.join(settings.RESULTS_FOLDER, output_filename)
    print(output_path)
    print(output_filename)
    wb.remove(wb.active)
    wb.save(output_path)
    return output_filename

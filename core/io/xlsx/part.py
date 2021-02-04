from typing import Iterator
import os

import openpyxl
import openpyxl.utils.exceptions

from config import settings
from core.io.xlsx.utilities import xlsx
from core.models import part as part_
from core.utilities import value


def read_parts_iter(xlsx_file) -> (Iterator, int):
    try:
        wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    except FileNotFoundError:
        print('Файл не найден: ', xlsx_file)
        return None, 0
    except openpyxl.utils.exceptions.InvalidFileException:
        print('Некорректный формат файла: ', xlsx_file)
        return None, 0

    ws = wb.worksheets[0]

    if ws is None:
        return None, 0

    start_row = 2 if ws.max_row > 1 and (value.has_cyrillic(ws[1][0].value) or not ws[1][0].value) else 1

    iter_parts = _build_parts_iterator(ws, start_row)
    parts_count = ws.max_row - start_row + 1
    return iter_parts, parts_count


def start_write_parts(time_moment, table_name):
    headers = ['Артикул', 'Бренд', 'Наименование', time_moment]
    return xlsx.start_write_xlsx(headers, table_name)


def save_temp_parts(wb, out_name):
    xlsx_file_path = os.path.join(settings.TEMP_XLSX_DIRECTORY, out_name)
    wb.save(xlsx_file_path)


def write_parts_to_xlsx(ws, parts: list):
    rows = (_get_row_from_part(part) for part in parts)
    xlsx.add_rows_xlsx(ws, rows)


def _build_parts_iterator(ws, start_row: int) -> Iterator:
    return (part_.Part(_prepare_part_number(str(row[0].value)), str(row[1].value))
         for row in ws.iter_rows(start_row, ws.max_row, 1, 3) if row[0].value and row[1].value)


def _get_row_from_part(part):
    return [part.number, part.model, part.title, part.price]


def _prepare_part_number(number):
    return number.split('#')[0]

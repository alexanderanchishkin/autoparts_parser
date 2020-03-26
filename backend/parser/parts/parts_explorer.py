import os.path
import re
import settings
import time

from flask import current_app

from openpyxl import load_workbook, Workbook
from backend.parser.parts.part import Part


def has_cyrillic(text):
    return bool(re.search('[а-яА-Я]', text))


def read_parts_from_xlsx(xlsx_file, max_amount=int(1e+6)):
    wb = load_workbook(xlsx_file)
    ws = wb.worksheets[0]
    start_row = 2 if not ws[1][0].value or has_cyrillic(ws[1][0].value) else 1
    parts = [Part(ws[row][0].value, ws[row][1].value) for row in range(start_row, min(ws.max_row + 1, max_amount + 1))]
    return parts


def write_column_to_xlsx(xlsx_file, xlsx_table, time_moment):
    xlsx_file_path = settings.OUTPUT_DIRECTORY + xlsx_file

    if not os.path.exists(xlsx_file_path):
        create_xlsx(xlsx_file_path, xlsx_table)

    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    add_new_column(ws, time_moment)
    while True:
        try:
            wb.save(xlsx_file_path)
            break
        except PermissionError:
            print('Закройте файл ', xlsx_file_path)
            time.sleep(3)


def write_parts_to_xlsx(xlsx_file, xlsx_table, parts, time_moment):
    xlsx_file_path = settings.OUTPUT_DIRECTORY + xlsx_file

    if not os.path.exists(xlsx_file_path):
        create_xlsx(xlsx_file_path, xlsx_table)
    wb = load_workbook(xlsx_file_path)
    ws = wb.active

    last_column = get_column_by_index(ws.max_column)

    for row in range(2, ws.max_row + 1):
        number = ws[row][0].value
        if number not in parts:
            continue
        part = parts[number]
        cell = last_column + str(row)

        ws[cell] = part.price
        parts.pop(number)

    for part in parts.values():
        write_new_part(ws, part, last_column)

    while True:
        try:
            wb.save(xlsx_file_path)
            break
        except PermissionError:
            print('Закройте файл ', xlsx_file_path)
            time.sleep(3)


def write_new_part(table, part, last_column):
    row = table.max_row + 1
    table['A' + str(row)] = str(part.number)
    table['B' + str(row)] = part.model
    table['C' + str(row)] = part.title
    table[last_column + str(row)] = part.price


def add_new_column(table, time_moment):
    new_column_name = get_column_by_index(table.max_column + 1) + '1'
    table[new_column_name] = time_moment


def create_xlsx(xlsx_file_path, xlsx_table):
    wb = Workbook()
    ws = wb.active
    ws.title = xlsx_table

    column_names = ['Артикул', 'Бренд', 'Наименование']
    add_header_row(ws, column_names)

    wb.save(xlsx_file_path)


def add_header_row(table, column_names):
    for index in range(len(column_names)):
        cell_name = get_column_by_index(index + 1) + '1'
        table[cell_name] = column_names[index]


def get_column_by_index(index):  # col is 1 based
    excel_column = str()
    div = index
    while div:
        (div, mod) = divmod(div-1, 26)  # will return (x, 0 .. 25)
        excel_column = chr(mod + 65) + excel_column

    return excel_column


def merge_files(files, out_name):
    wb = Workbook()
    folder = 'results'
    for file in files:
        filename = os.path.join(folder, file)
        file_wb = load_workbook(filename)
        file_ws = file_wb.active
        wb.create_sheet(file_ws.title)
        for row in file_ws.rows:
            wb[file_ws.title].append((ceil.value for ceil in row))
    output_filename = f"{out_name}_{settings.TIME_MOMENT.replace(' ', '').replace('.', '').replace(':', '')}.xlsx"
    output_path = os.path.join(folder, output_filename)
    print(output_path)
    print(output_filename)
    wb.remove(wb.active)
    wb.save(output_path)
    return output_filename
